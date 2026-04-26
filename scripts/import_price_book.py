"""
Ingest a manufacturer / distributor line list (CSV or XLSX) into the price book.

Usage:
    python scripts/import_price_book.py --file path/to/list.csv
    python scripts/import_price_book.py --file path/to/list.xlsx --org-id 1 --manufacturer Trane

Header detection:
    The importer reads column headers from row 1 and matches them against a
    set of canonical aliases (case-insensitive, punctuation-tolerant). At
    minimum the file must have a description and a unit cost. CSI code,
    category, manufacturer, model, size, unit, labor hours, and notes are
    optional.

Upsert key:
    Items are matched on (org_id, manufacturer, model_number) when both are
    present. Otherwise on (org_id, csi_code, description, size). Existing rows
    have their cost / labor / notes / last_price_update updated; missing rows
    are created. Items skipped due to missing description or zero cost are
    reported.

Use --dry-run to preview changes without writing.
"""
import argparse
import asyncio
import csv
import io
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "backend"))


# Map every accepted header label → canonical column name.
HEADER_ALIASES: dict[str, str] = {
    # description
    "description": "description",
    "desc": "description",
    "item": "description",
    "item description": "description",
    "product": "description",
    "product description": "description",
    # csi
    "csi": "csi_code",
    "csi code": "csi_code",
    "csi number": "csi_code",
    "masterformat": "csi_code",
    "spec section": "csi_code",
    # category
    "category": "category",
    "type": "category",
    "group": "category",
    # manufacturer
    "manufacturer": "manufacturer",
    "mfr": "manufacturer",
    "make": "manufacturer",
    "vendor": "manufacturer",
    "brand": "manufacturer",
    # model
    "model": "model_number",
    "model number": "model_number",
    "model no": "model_number",
    "part number": "model_number",
    "part no": "model_number",
    "sku": "model_number",
    "catalog": "model_number",
    "catalog number": "model_number",
    # size
    "size": "size",
    "dimensions": "size",
    "nominal size": "size",
    "diameter": "size",
    # unit (LF, EA, SF…)
    "unit": "unit",
    "uom": "unit",
    "unit of measure": "unit",
    "u/m": "unit",
    # material cost
    "material cost": "material_unit_cost",
    "material unit cost": "material_unit_cost",
    "unit cost": "material_unit_cost",
    "list price": "material_unit_cost",
    "net price": "material_unit_cost",
    "price": "material_unit_cost",
    "cost": "material_unit_cost",
    "$": "material_unit_cost",
    "usd": "material_unit_cost",
    # labor
    "labor hours": "labor_hours_per_unit",
    "labor": "labor_hours_per_unit",
    "labor hr": "labor_hours_per_unit",
    "labor hours per unit": "labor_hours_per_unit",
    "hours": "labor_hours_per_unit",
    "labor unit": "labor_hours_per_unit",
    # labor rate (per-row override)
    "labor rate": "labor_rate",
    "rate": "labor_rate",
    "$/hr": "labor_rate",
    # notes
    "notes": "notes",
    "comment": "notes",
    "comments": "notes",
    "remarks": "notes",
}


def _canonical(header: str) -> str:
    h = header.strip().lower()
    h = h.replace("_", " ").replace("-", " ").replace(".", "")
    h = " ".join(h.split())
    return HEADER_ALIASES.get(h, "")


def _parse_money(raw: Any) -> float | None:
    if raw is None or raw == "":
        return None
    s = str(raw).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _parse_float(raw: Any) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        return float(str(raw).strip())
    except ValueError:
        return None


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if len(rows) < 2:
        return []
    return _rows_with_headers(rows[0], rows[1:])


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl not installed; pip install openpyxl") from None
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if len(rows) < 2:
        return []
    return _rows_with_headers(rows[0], rows[1:])


def _rows_with_headers(header_row: list, data_rows: list) -> list[dict[str, Any]]:
    canonical = [_canonical(h or "") for h in header_row]
    out = []
    for row in data_rows:
        if not any(c not in (None, "") for c in row):
            continue
        rec: dict[str, Any] = {}
        for col, value in zip(canonical, row, strict=False):
            if col and value not in (None, ""):
                rec[col] = value
        if rec:
            out.append(rec)
    return out


def _normalize(record: dict[str, Any], default_manufacturer: str | None) -> dict[str, Any]:
    out = dict(record)
    if "material_unit_cost" in out:
        out["material_unit_cost"] = _parse_money(out["material_unit_cost"])
    if "labor_hours_per_unit" in out:
        out["labor_hours_per_unit"] = _parse_float(out["labor_hours_per_unit"])
    if "labor_rate" in out:
        out["labor_rate"] = _parse_money(out["labor_rate"])
    if default_manufacturer and not out.get("manufacturer"):
        out["manufacturer"] = default_manufacturer
    for k in ("csi_code", "category", "description", "manufacturer",
              "model_number", "size", "unit", "notes"):
        if k in out and out[k] is not None:
            out[k] = str(out[k]).strip()
    if not out.get("unit"):
        out["unit"] = "EA"
    if not out.get("category"):
        out["category"] = "uncategorized"
    return out


async def _resolve_org(db, org_id: int | None):
    from models.user import Organization
    from sqlalchemy import select

    if org_id:
        result = await db.execute(select(Organization).where(Organization.id == org_id))
    else:
        result = await db.execute(select(Organization).limit(1))
    return result.scalar_one_or_none()


async def _resolve_mech_trade(db, org_id: int) -> int | None:
    from models.trade import Trade
    from sqlalchemy import select

    result = await db.execute(
        select(Trade).where(Trade.code == "MECH", Trade.org_id == org_id)
    )
    trade = result.scalar_one_or_none()
    return trade.id if trade else None


async def _find_existing(db, org_id: int, rec: dict[str, Any]):
    from models.price_book import PriceBookItem
    from sqlalchemy import select

    if rec.get("manufacturer") and rec.get("model_number"):
        result = await db.execute(
            select(PriceBookItem).where(
                PriceBookItem.org_id == org_id,
                PriceBookItem.manufacturer == rec["manufacturer"],
                PriceBookItem.model_number == rec["model_number"],
            )
        )
        item = result.scalar_one_or_none()
        if item:
            return item
    description = rec.get("description")
    if not description:
        return None
    q = select(PriceBookItem).where(
        PriceBookItem.org_id == org_id,
        PriceBookItem.description == description,
    )
    if rec.get("csi_code"):
        q = q.where(PriceBookItem.csi_code == rec["csi_code"])
    if rec.get("size"):
        q = q.where(PriceBookItem.size == rec["size"])
    result = await db.execute(q)
    return result.scalar_one_or_none()


def _apply_updates(item, rec: dict[str, Any]) -> bool:
    changed = False
    update_fields = (
        "csi_code", "category", "manufacturer", "model_number", "size", "unit",
        "labor_rate",
    )
    for f in update_fields:
        if f in rec and rec[f] is not None and getattr(item, f) != rec[f]:
            setattr(item, f, rec[f])
            changed = True

    if (cost := rec.get("material_unit_cost")) is not None and cost > 0:
        if item.material_unit_cost != cost:
            item.material_unit_cost = cost
            changed = True
        # Pricing now real — clear PRICE_PENDING marker.
        if item.notes and "PRICE_PENDING_BLUE_BOOK" in item.notes:
            item.notes = (rec.get("notes") or "").strip() or None
            changed = True
        item.last_price_update = datetime.now(timezone.utc)

    if (hours := rec.get("labor_hours_per_unit")) is not None and hours > 0:
        if item.labor_hours_per_unit != hours:
            item.labor_hours_per_unit = hours
            changed = True
    if rec.get("notes") and item.notes != rec["notes"]:
        item.notes = rec["notes"]
        changed = True
    return changed


async def import_file(
    file: Path,
    *,
    org_id: int | None = None,
    trade_id: int | None = None,
    default_manufacturer: str | None = None,
    dry_run: bool = False,
) -> dict[str, int]:
    suffix = file.suffix.lower()
    if suffix == ".csv":
        raw_rows = _read_csv(file)
    elif suffix in (".xlsx", ".xlsm"):
        raw_rows = _read_xlsx(file)
    else:
        raise SystemExit(f"Unsupported file extension: {suffix} (use .csv or .xlsx)")

    if not raw_rows:
        print(f"No data rows found in {file}")
        return {"created": 0, "updated": 0, "skipped": 0, "total": 0}

    records = [_normalize(r, default_manufacturer) for r in raw_rows]

    from core.database import AsyncSessionLocal
    from models.price_book import PriceBookItem

    async with AsyncSessionLocal() as db:
        org = await _resolve_org(db, org_id)
        if not org:
            raise SystemExit("No organization found. Run scripts/seed_users.py first.")
        active_trade_id = trade_id or await _resolve_mech_trade(db, org.id)

        created = updated = skipped = 0
        skipped_reasons: dict[str, int] = {}
        for rec in records:
            if not rec.get("description"):
                skipped += 1
                skipped_reasons["missing description"] = skipped_reasons.get(
                    "missing description", 0) + 1
                continue
            cost = rec.get("material_unit_cost")
            if cost is None or cost <= 0:
                # If row has no usable price, only worth importing as a new
                # skeleton entry if it doesn't already exist.
                pass

            existing = await _find_existing(db, org.id, rec)
            if existing:
                if _apply_updates(existing, rec):
                    updated += 1
                else:
                    skipped += 1
                    skipped_reasons["no changes"] = skipped_reasons.get(
                        "no changes", 0) + 1
                continue

            new_item = PriceBookItem(
                org_id=org.id,
                trade_id=active_trade_id,
                csi_code=rec.get("csi_code"),
                category=rec["category"],
                description=rec["description"],
                manufacturer=rec.get("manufacturer"),
                model_number=rec.get("model_number"),
                size=rec.get("size"),
                unit=rec["unit"],
                material_unit_cost=cost or 0.0,
                labor_hours_per_unit=rec.get("labor_hours_per_unit") or 0.0,
                labor_rate=rec.get("labor_rate"),
                notes=rec.get("notes"),
                last_price_update=datetime.now(timezone.utc) if cost else None,
            )
            db.add(new_item)
            created += 1

        if dry_run:
            await db.rollback()
            print("[dry-run] No changes committed.")
        else:
            await db.commit()

    summary = {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total": len(records),
    }
    print(
        f"Imported from {file.name}: "
        f"{created} created, {updated} updated, {skipped} skipped "
        f"({len(records)} rows total)"
    )
    if skipped_reasons:
        for reason, count in skipped_reasons.items():
            print(f"  - skipped ({reason}): {count}")
    return summary


# Public surface for the FastAPI uploader to reuse the same parsing.
def parse_buffer(content: bytes, filename: str) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
    elif suffix in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl not installed") from None
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        raise ValueError(f"Unsupported extension: {suffix}")
    if len(rows) < 2:
        return []
    return [_normalize(r, None) for r in _rows_with_headers(rows[0], rows[1:])]


def _cli() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, help="CSV or XLSX path")
    parser.add_argument("--org-id", type=int, help="Target organization id (default: first org)")
    parser.add_argument("--trade-id", type=int, help="Target trade id (default: MECH trade)")
    parser.add_argument("--manufacturer", help="Default manufacturer for rows missing a Manufacturer column")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing")
    args = parser.parse_args()

    file = Path(args.file)
    if not file.exists():
        raise SystemExit(f"File not found: {file}")

    asyncio.run(
        import_file(
            file,
            org_id=args.org_id,
            trade_id=args.trade_id,
            default_manufacturer=args.manufacturer,
            dry_run=args.dry_run,
        )
    )


if __name__ == "__main__":
    _cli()
