import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_to_excel(bid, line_items) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Bid v{bid.version}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a3a5c")
    total_fill = PatternFill("solid", fgColor="e8f0fe")
    currency_fmt = '#,##0.00'
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"BID SUMMARY — {bid.name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    # Headers
    headers = ["#", "Description", "Category", "System", "Qty", "Unit", "Unit Material", "Unit Labor Hrs", "Material Total", "Labor Total", "Line Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_num, item in enumerate(line_items, start=4):
        ws.cell(row=row_num, column=1, value=row_num - 3)
        ws.cell(row=row_num, column=2, value=item.description)
        ws.cell(row=row_num, column=3, value=item.category)
        ws.cell(row=row_num, column=4, value=item.system)
        ws.cell(row=row_num, column=5, value=item.quantity)
        ws.cell(row=row_num, column=6, value=item.unit)
        ws.cell(row=row_num, column=7, value=item.unit_material_cost).number_format = currency_fmt
        ws.cell(row=row_num, column=8, value=item.unit_labor_hours)
        ws.cell(row=row_num, column=9, value=item.material_total).number_format = currency_fmt
        ws.cell(row=row_num, column=10, value=item.labor_total).number_format = currency_fmt
        ws.cell(row=row_num, column=11, value=item.line_total).number_format = currency_fmt

    last_data = len(line_items) + 3

    # Totals section
    totals_start = last_data + 2
    totals = [
        ("Material Cost", bid.total_material_cost),
        ("Material Markup", bid.total_material_markup),
        ("Labor Cost", bid.total_labor_cost),
        ("Labor Burden", bid.total_burden),
        ("Overhead", bid.total_overhead),
        ("Subtotal", bid.subtotal),
        ("Contingency", bid.contingency),
        ("Bond", bid.bond),
        ("Permit & Fees", bid.permit),
        ("Profit", bid.profit),
        ("GRAND TOTAL", bid.grand_total),
    ]
    for i, (label, value) in enumerate(totals):
        row = totals_start + i
        label_cell = ws.cell(row=row, column=9, value=label)
        value_cell = ws.cell(row=row, column=11, value=value)
        value_cell.number_format = currency_fmt
        if label == "GRAND TOTAL":
            label_cell.font = Font(bold=True, size=12)
            value_cell.font = Font(bold=True, size=12)
            value_cell.fill = PatternFill("solid", fgColor="ffd700")

    # Column widths
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    for col in range(5, 12):
        ws.column_dimensions[get_column_letter(col)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
